from openpyxl import Workbook, load_workbook
from googletrans import Translator
from tqdm import tqdm

def new_trans(file_excel):
    wb = load_workbook(file_excel)
    sheet = wb.worksheets[0]
    translator = Translator()

    updated_rows = []

    for row in tqdm(sheet.iter_rows(min_row=2, values_only=True)):
        value_column_3 = row[2]  # Access value in column 3 (index 2)
        print("origin: ",value_column_3)
        # Perform your processing and store the result
        translate = translator.translate(value_column_3, dest='en')
        print("result : ",translate.text)
        updated_rows.append(translate.text)
        # Write the result in column 4 (index 3)
    for idx, updated_row in enumerate(updated_rows, start=1):  # Start from row 2
        sheet[f"D{idx}"].value = updated_row  # Write the updated value in column 4
    print(updated_rows)
    wb.save("output.xlsx")

if __name__ == "__main__":
    pos_col = (ord('c') - 96)
    new_trans("ALL MACHINE PLATING IO LIST ADDRESS.xlsx")
