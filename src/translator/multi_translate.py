from openpyxl import Workbook, load_workbook
from concurrent.futures import ThreadPoolExecutor
from deep_translator import GoogleTranslator
from tqdm import tqdm

wb = load_workbook("ALL MACHINE PLATING IO LIST ADDRESS.xlsx")
sheet = wb.active

temp_list = []
for row in tqdm(sheet.iter_rows(min_row=2, values_only=True)):
    value = row[2]
    temp_list.append(value)
print(temp_list)
