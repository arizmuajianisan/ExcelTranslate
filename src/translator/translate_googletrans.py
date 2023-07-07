from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import concurrent.futures
from googletrans import Translator


def translate_batch(values):
    translator = Translator()
    translated = []
    retries = 3  # Maximum number of retries

    for _ in range(retries):
        try:
            translated = [translator.translate(value, dest="en") for value in values]
            break
        except Exception as e:
            print(f"Translation failed. Retrying... ({e})")

    return [result.text for result in translated]

def parameter(self):
    self.file_want_to_translate = file_want_to_translate
    self.output_file = output_file
    self.sheet_selector = sheet_selector
    self.column_selector = column_selector


def translate_one_sheet(file_excel, name_file, position, row_selector):
    wb = load_workbook(file_excel)  # pos=-1 to match index of worksheet, start from 0
    sheet = wb.worksheets[position - 1]

    values_to_translate = [row[row_selector - 1].value for row in sheet.iter_rows(min_row=1)]
    translated_values = []

    with concurrent.futures.ThreadPoolExecutor() as executor:
        batch_size = 100  # Number of values to translate in each batch
        for i in tqdm(range(0, len(values_to_translate), batch_size)):
            batch = values_to_translate[i : i + batch_size]
            results = executor.submit(translate_batch, batch)
            translated = results.result()
            translated_values.extend(translated)

    for idx, row in tqdm(
        enumerate(sheet.iter_rows(min_col=1, max_col=4, min_row=1), start=1),
        desc="Updating rows",
    ):
        row[row_selector + 1 - 1].value = translated_values[idx - 1]

    wb.save(f"{name_file}.xlsx")


def translate_all_sheet(file_excel, name_file):
    wb = load_workbook(file_excel)

    for sheet in wb.worksheets[0:]:
        values_to_translate = [row[2].value for row in sheet.iter_rows(min_row=1)]
        translated_values = []

        with concurrent.futures.ThreadPoolExecutor() as executor:
            batch_size = 100  # Number of values to translate in each batch
            for i in tqdm(range(0, len(values_to_translate), batch_size)):
                batch = values_to_translate[i : i + batch_size]
                results = executor.submit(translate_batch, batch)
                translated = results.result()
                translated_values.extend(translated)

        for idx, row in tqdm(
            enumerate(sheet.iter_rows(min_col=1, max_col=4, min_row=1), start=1),
            desc="Updating rows",
        ):
            row[3].value = translated_values[idx - 1]

    wb.save(f"{name_file}.xlsx")


if __name__ == "__main__":
    print("*** Ini akan translate semua column yang di select dan return di +1 column ***\n")
    type_trans = input("Mode translate? options: one - all >>> ")
    match type_trans:
        case "one":
            file_want_to_translate = input("Enter excel file to translate >>> ")
            name_file = input("Expected output name file >>> ")
            pages = int(input("Enter sheet >>> "))
            row_selector = ord((input("Enter column to translate, ie: A,B,C >>> ").lower())) - 96
            print(row_selector)
            translate_one_sheet(f"{file_want_to_translate}.xlsx", name_file, pages, row_selector)
        case "all":
            print("BELOM KELAR")
            # name_file = input("Expected output name file >>> ")
            # translate_all_sheet("ALL MACHINE PLATING IO LIST ADDRESS.xlsx", name_file)
        case _:
            print("Wrong option")
